# ✅ ESTA LINHA TEM QUE SER A PRIMEIRA EXECUTÁVEL!
import streamlit as st
st.set_page_config(page_title="Confronto XML x Invoice", layout="wide")

import os
import pandas as pd
import io
import requests
from datetime import datetime, timedelta
from utils import parser_xml, parser_invoice, comparador
import openpyxl
from openpyxl.styles import PatternFill

st.title("📦 Confronto de XML da NF-e com Invoice (CI)")

# ---- Função para destacar duplicados no Excel (versão manual, qualquer openpyxl) ----
def destacar_duplicados_manual(buffer, colunas_destacar=[2, 8], cor_hex="FFFF00"):
    wb = openpyxl.load_workbook(buffer)
    ws = wb.active

    for col in colunas_destacar:
        col_letter = openpyxl.utils.get_column_letter(col)
        # Pega todos os valores da coluna, desconsiderando o cabeçalho
        valores = [ws[f"{col_letter}{row}"].value for row in range(2, ws.max_row + 1)]
        duplicados = set([v for v in valores if valores.count(v) > 1 and v is not None])
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            if cell.value in duplicados:
                cell.fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ---- Box da cotação do dólar no topo direito ----
def get_cotacao_usd():
    try:
        hoje = datetime.now()
        ontem = hoje - timedelta(days=1)
        url = "https://economia.awesomeapi.com.br/json/daily/USD-BRL/2"
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        cotacoes = r.json()
        cotacoes_ordenadas = sorted(
            cotacoes,
            key=lambda x: x['timestamp'],
            reverse=True
        )
        cotacao_hoje = cotacoes_ordenadas[0]
        cotacao_ontem = cotacoes_ordenadas[1]
        dolar_hoje = float(cotacao_hoje['bid'])
        data_hoje_str = datetime.fromtimestamp(int(cotacao_hoje['timestamp'])).strftime('%d/%m/%Y')
        dolar_ontem = float(cotacao_ontem['bid'])
        data_ontem_str = datetime.fromtimestamp(int(cotacao_ontem['timestamp'])).strftime('%d/%m/%Y')
        return {
            "hoje": {"valor": dolar_hoje, "data": data_hoje_str},
            "ontem": {"valor": dolar_ontem, "data": data_ontem_str}
        }
    except Exception as e:
        return None

cotacoes = get_cotacao_usd()
st.markdown(" ")
_, col_cotacao = st.columns([7, 1])
if cotacoes:
    st.markdown(
        f"""
        <div style='display:flex;flex-direction:column;align-items:end; margin-top:-30px; margin-bottom:10px;'>
          <div style='background:#232931;padding:18px 32px;border-radius:18px;box-shadow:0 2px 8px #00000015;min-width:260px;'>
            <div style='font-size:1.7rem;font-weight:700;color:#3be180;margin-bottom:5px;'>
                <span style="font-size:2rem;">$</span> Dólar Comercial
            </div>
            <div style='font-size:1rem;color:#f5f5f5;margin-bottom:7px;'>
                Cotação hoje <b>({cotacoes['hoje']['data']})</b><br>
                <span style='font-size:1.3rem;color:#fff;font-weight:700;'>R$ {cotacoes['hoje']['valor']:.4f}</span>
            </div>
            <div style='font-size:1rem;color:#f5f5f5;'>
                Cotação ontem <b>({cotacoes['ontem']['data']})</b><br>
                <span style='font-size:1.3rem;color:#fff;font-weight:700;'>R$ {cotacoes['ontem']['valor']:.4f}</span>
            </div>
          </div>
        </div>
        """,
        unsafe_allow_html=True
    )
else:
    st.error("Não foi possível obter a cotação do dólar.")

# --- Upload dos arquivos ---
st.header("📁 Upload de Arquivos")

modelo_path = os.path.join("utils", "modelo_invoice.xlsx")
with st.expander("📥 Baixar modelo da planilha Invoice"):
    try:
        with open(modelo_path, "rb") as f:
            st.download_button(
                label="📥 Clique aqui para baixar o modelo (.xlsx)",
                data=f,
                file_name="modelo_invoice.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        st.info("Use esse modelo para colar os dados da sua Invoice no formato esperado.")
    except FileNotFoundError:
        st.error("Arquivo 'modelo_invoice.xlsx' não encontrado em `utils/`.")

col1, col2 = st.columns(2)

with col1:
    xml_file = st.file_uploader("📤 Enviar XML da NF-e", type=["xml"], key="xml")

with col2:
    invoice_file = st.file_uploader("📤 Enviar Planilha da Invoice (aba 'CI')", type=["xlsx", "xls"], key="invoice")

# --- Checkbox Grade ---
grade_mode = st.checkbox("Usar cálculo por Grade (total pares = quantidade × caixas por tamanho)")

# --- NOVO BLOCO: Conversão de Moeda ---
st.markdown("### 💵 Opções de Conversão de Moeda")

col_a, col_b, col_c, col_d = st.columns([1,1,2,2])
with col_a:
    xml_em_dolar = st.checkbox("XML em dólar")
with col_b:
    invoice_em_dolar = st.checkbox("Invoice em dólar")
with col_c:
    usar_cotacao_auto = st.checkbox("Usar cotação automática (dia anterior)", value=True)
with col_d:
    cotacao_manual = st.number_input("Cotação manual (opcional)", value=0.0, format="%.4f")

if xml_em_dolar and invoice_em_dolar:
    st.error("Selecione apenas um dos campos como 'em dólar' para realizar a conversão.")
    st.stop()

cotacao_dolar = None
if usar_cotacao_auto or cotacao_manual == 0:
    cotacao_dolar = cotacoes['ontem']['valor'] if cotacoes else 1.0
else:
    cotacao_dolar = cotacao_manual if cotacao_manual > 0 else (cotacoes['ontem']['valor'] if cotacoes else 1.0)

def converter_para_reais(itens, campos, cotacao):
    for item in itens:
        for campo in campos:
            if campo in item:
                try:
                    valor = float(str(item[campo]).replace(",", "."))
                    item[campo] = f"{valor * cotacao:.10f}".rstrip("0").rstrip(".")
                except Exception:
                    pass
    return itens

# --- Processamento após upload ---
if xml_file and invoice_file:
    st.success("✅ Arquivos carregados com sucesso.")

    # Parse dos dados
    st.header("🔍 Prévia dos Dados Carregados")

    with st.spinner("Lendo XML..."):
        dados_xml, resumo_xml = parser_xml.processar(xml_file)

    with st.spinner("Lendo Invoice..."):
        dados_invoice, resumo_invoice = parser_invoice.processar(invoice_file, usar_grade=grade_mode)

    if "erro" in resumo_invoice:
        st.error(f"Erro ao processar Invoice: {resumo_invoice['erro']}")
        st.stop()

    # -- CONVERSÃO DE MOEDA ANTES DO CONFRONTO --
    campos_monetarios_xml = ["preco unit xml", "valor total xml"]
    campos_monetarios_invoice = ["preco unit invoice", "valor total invoice"]

    if xml_em_dolar:
        dados_xml = converter_para_reais(dados_xml, campos_monetarios_xml, cotacao_dolar)
        if "valor total xml" in resumo_xml:
            resumo_xml["valor total xml"] = round(float(resumo_xml["valor total xml"]) * cotacao_dolar, 10)
    if invoice_em_dolar:
        dados_invoice = converter_para_reais(dados_invoice, campos_monetarios_invoice, cotacao_dolar)
        if "valor total nota" in resumo_invoice:
            resumo_invoice["valor total nota"] = round(float(resumo_invoice["valor total nota"]) * cotacao_dolar, 10)

    # Mostrar resumos
    st.subheader("📑 Resumo do XML")
    st.json(resumo_xml, expanded=False)

    st.subheader("📑 Resumo da Invoice (CI)")
    st.json(resumo_invoice, expanded=False)

    # Mostrar agrupamento dos itens do XML e da Invoice
    st.subheader("🔎 Itens do XML (após conversão, se houver)")
    st.dataframe(pd.DataFrame(dados_xml), use_container_width=True)

    st.subheader("🔎 Itens da Invoice (após conversão, se houver)")
    st.dataframe(pd.DataFrame(dados_invoice), use_container_width=True)

    # Confronto
    if st.button("🚨 Confrontar XML x Invoice"):
        with st.spinner("Comparando os dados..."):
            resultado = comparador.confrontar(dados_xml, dados_invoice)
            st.session_state["resultado"] = resultado
            st.session_state["mostrar_erros"] = False

    # Exibir resultado do confronto (se houver)
    if "resultado" in st.session_state:
        resultado = st.session_state["resultado"]

        st.subheader("📊 Resultado do Confronto (Agrupado)")
        st.dataframe(resultado, use_container_width=True)

        # --- RESUMO DE ERROS ---
        erros = resultado[
            (resultado["verificação ref"] != "✅ OK") |
            (resultado["verificação ncm"] != "✅ OK") |
            (resultado["verificação cor"] != "✅ OK") |
            (resultado["verificação total pares"] != "✅ OK") |
            (resultado["verificação preco unit"] != "✅ OK") |
            (resultado["verificação valor total"] != "✅ OK")
        ]
        st.session_state["erros"] = erros
        qtd_erros = len(erros)

        st.markdown("### ❗ Resumo de Divergências")
        col1, col2 = st.columns(2)

        with col1:
            st.metric("Total de Itens com Erros", qtd_erros)

        with col2:
            if st.button("🔍 Exibir Apenas os Erros"):
                st.session_state["mostrar_erros"] = True

        if st.session_state.get("mostrar_erros", False):
            st.subheader("🧯 Itens com Divergência")
            st.dataframe(erros, use_container_width=True)

        if not erros.empty:
            buffer = io.BytesIO()
            erros.to_excel(buffer, index=False)
            buffer.seek(0)
            # Aplica destaque de duplicados manualmente em ref xml (B) e cor xml (H)
            buffer = destacar_duplicados_manual(buffer, colunas_destacar=[2, 8])
            st.download_button(
                label="📥 Baixar Erros em Excel (com duplicados destacados)",
                data=buffer,
                file_name="erros_confronto.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    st.markdown("---")
    if st.button("🔄 Iniciar Nova Confrontação"):
        st.session_state.clear()
        st.rerun()

else:
    st.warning("⚠️ Envie os dois arquivos para iniciar.")
